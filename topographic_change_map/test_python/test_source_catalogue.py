from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parents[1]
sys.path.insert(0, str(ROOT / "python"))

from build_source_spreadsheet import build_workbook, validate, write_csv  # noqa: E402
from source_catalogue_data import AUTH_PROFILES, NEXT_STEPS, SOURCES  # noqa: E402


def test_catalogue_is_large_normalized_and_actionable() -> None:
    assert len(SOURCES) >= 60
    assert len({item["id"] for item in SOURCES}) == len(SOURCES)
    assert all(item["auth_profile"] in AUTH_PROFILES for item in SOURCES)
    assert all(item["source_url"].startswith("https://") for item in SOURCES)
    assert all(item["next_action"] for item in SOURCES if item["priority"] in {"P0", "P1"})
    assert len(NEXT_STEPS) >= 10


def test_every_auth_profile_is_used() -> None:
    used = {item["auth_profile"] for item in SOURCES}
    assert used == set(AUTH_PROFILES)


def test_workbook_round_trip_and_validation(tmp_path: Path) -> None:
    xlsx = tmp_path / "catalogue.xlsx"
    csv = tmp_path / "catalogue.csv"
    build_workbook(xlsx)
    write_csv(csv)
    report = validate(xlsx, csv)
    assert report["passed"], report

    workbook = load_workbook(xlsx)
    assert workbook.sheetnames == [
        "Executive Summary",
        "Prioritized Next Steps",
        "Source Catalogue",
        "Authentication Guide",
        "Nepal Pair Evidence",
        "Field Definitions",
        "References",
    ]
    assert workbook["Source Catalogue"].max_row == len(SOURCES) + 4
    assert workbook["Source Catalogue"].freeze_panes == "F5"
    assert len(workbook["References"].tables) == 1

#!/usr/bin/env python3
"""Build and validate the Nepal flood data-source catalogue workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from source_catalogue_data import (
    AUTH_PROFILES,
    CURRENT_STATE,
    FIELD_DEFINITIONS,
    NEXT_STEPS,
    SOURCES,
)


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "research" / "Nepal_Flood_Topographic_Data_Source_Catalogue.xlsx"
DEFAULT_CSV = ROOT / "research" / "imagery-source-catalogue.csv"
DEFAULT_REPORT = ROOT / "research" / "source-catalogue-validation.json"

NAVY = "071827"
NAVY_2 = "0E2B3C"
CYAN = "47D7F2"
WHITE = "FFFFFF"
TEXT = "102A3A"
MUTED = "5C7080"
PALE_BLUE = "EAF7FB"
PALE_GREY = "F3F6F8"
GREEN = "D9EAD3"
YELLOW = "FFF2CC"
ORANGE = "FCE5CD"
RED = "F4CCCC"
PURPLE = "EADCF8"
BLUE = "D9EAF7"
GREY = "E7E6E6"
THIN = Side(style="thin", color="CBD8DF")

SOURCE_COLUMNS = [
    ("id", "ID"),
    ("priority", "Priority"),
    ("category", "Category"),
    ("provider", "Provider / programme"),
    ("sensor_product", "Sensor / product"),
    ("nepal_candidate", "Nepal candidate / coverage"),
    ("epoch_role", "Epoch role"),
    ("resolution", "Resolution"),
    ("height_method", "Height / change method"),
    ("geometry", "Geometry / camera / phase"),
    ("product_needed", "Product needed"),
    ("what_exists", "What exists"),
    ("visited_status", "Research / visit status"),
    ("provided_to_project", "Provided to this project"),
    ("usefulness", "Usefulness"),
    ("verdict", "Current verdict"),
    ("principal_blocker", "Principal blocker"),
    ("access_tier", "Access tier"),
    ("account_required", "Account required"),
    ("signup_requirements", "Signup / identity requirements"),
    ("verification_mfa", "Verification / MFA"),
    ("eligibility_approval", "Eligibility / approval"),
    ("contract_licence_payment", "Contract / licence / payment"),
    ("api_authentication", "API authentication"),
    ("delivery", "Delivery"),
    ("current_account_status", "Current account status"),
    ("current_entitlement", "Current entitlement"),
    ("next_action", "Next action"),
    ("source_url", "Primary source URL"),
    ("access_url", "Access / signup URL"),
    ("auth_docs_url", "Authentication docs URL"),
    ("confidence", "Confidence"),
    ("verified_date", "Verified date"),
    ("notes", "Notes"),
]


def _title(ws, title: str, subtitle: str, end_col: int = 8) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    cell = ws.cell(2, 1, subtitle)
    cell.fill = PatternFill("solid", fgColor=NAVY_2)
    cell.font = Font(name="Aptos", size=10, color=CYAN)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 31


def _section(cell, value: str) -> None:
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=NAVY_2)
    cell.font = Font(name="Aptos", bold=True, color=WHITE, size=11)
    cell.alignment = Alignment(vertical="center")


def _header(row) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=NAVY_2)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=CYAN))


def _body(ws, start_row: int, end_row: int, max_col: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=WHITE if row_idx % 2 else PALE_GREY)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = fill
            cell.font = Font(name="Aptos", color=TEXT, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN)


def _link(cell) -> None:
    if isinstance(cell.value, str) and cell.value.startswith("http"):
        cell.hyperlink = cell.value
        cell.font = Font(name="Aptos", color="0563C1", underline="single", size=9)


def _table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def _write_summary(wb: Workbook, generated_at: str) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    _title(
        ws,
        "Nepal 2026 flood — topographic data-source catalogue",
        f"Team decision workbook • generated {generated_at} • contains access instructions, never credentials or secrets",
        9,
    )
    ws.merge_cells("A4:I5")
    ws["A4"] = (
        "GOAL  Produce a scientifically defensible flood-related topographic-change map over the Bhote Koshi / "
        "Trishuli corridor by reconstructing compatible pre- and post-event elevation surfaces, co-registering "
        "them on stable terrain, propagating uncertainty, and retaining unsupported areas as unsupported."
    )
    ws["A4"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    ws["A4"].font = Font(name="Aptos", size=12, bold=True, color=TEXT)
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A4"].border = Border(left=Side(style="medium", color=CYAN))

    ws.merge_cells("A7:I8")
    ws["A7"] = (
        "CURRENT SCIENTIFIC LIMIT  The viewer's coloured layer is a relative post-event ortho-parallax height "
        "residual. It is not a true before/after elevation difference. A true change map still requires camera-"
        "bearing post-event data and a compatible pre-event surface, followed by independent alignment validation."
    )
    ws["A7"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["A7"].font = Font(name="Aptos", size=11, bold=True, color="7F6000")
    ws["A7"].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A7"].border = Border(left=Side(style="medium", color="E6B800"))

    ws.merge_cells("A10:D10")
    _section(ws["A10"], "Current programme state")
    row = 11
    for key, value in CURRENT_STATE:
        ws.cell(row, 1, key).font = Font(name="Aptos", bold=True, color=TEXT, size=9)
        ws.cell(row, 2, value)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        fill = PatternFill("solid", fgColor=WHITE if row % 2 else PALE_GREY)
        for col in range(1, 5):
            ws.cell(row, col).fill = fill
            ws.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, col).border = Border(bottom=THIN)
        row += 1

    ws.merge_cells("F10:I10")
    _section(ws["F10"], "Catalogue snapshot")
    metrics = [
        ("Sources / routes", len(SOURCES)),
        ("Immediate P0 routes", sum(s["priority"] == "P0" for s in SOURCES)),
        ("High-value P1 routes", sum(s["priority"] == "P1" for s in SOURCES)),
        ("Open / no-auth routes", sum(s["access_tier"] == "Open / no authentication" for s in SOURCES)),
        ("Commercially gated routes", sum("Commercial" in s["access_tier"] for s in SOURCES)),
        ("Institution/proposal/vetted routes", sum(any(k in s["access_tier"] for k in ("Institution", "proposal", "vetted")) for s in SOURCES)),
        ("Next-step decision gates", len(NEXT_STEPS)),
        ("Last evidence verification", max(s["verified_date"] for s in SOURCES)),
    ]
    for i, (key, value) in enumerate(metrics, 11):
        ws.cell(i, 6, key)
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=8)
        ws.cell(i, 9, value)
        for col in range(6, 10):
            ws.cell(i, col).fill = PatternFill("solid", fgColor=WHITE if i % 2 else PALE_GREY)
            ws.cell(i, col).border = Border(bottom=THIN)
            ws.cell(i, col).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(i, 6).font = Font(name="Aptos", bold=True, color=TEXT, size=9)
        ws.cell(i, 9).font = Font(name="Aptos", bold=True, color=NAVY, size=11)

    ws.merge_cells("F21:I21")
    _section(ws["F21"], "Sources by priority")
    counts = Counter(s["priority"] for s in SOURCES)
    for idx, priority in enumerate(("P0", "P1", "P2", "P3"), 22):
        ws.cell(idx, 6, priority)
        ws.cell(idx, 7, counts[priority])
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Catalogue routes"
    chart.y_axis.title = "Priority"
    chart.x_axis.title = "Count"
    chart.height = 6.5
    chart.width = 10.5
    chart.add_data(Reference(ws, min_col=7, min_row=21, max_row=25), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=6, min_row=22, max_row=25))
    chart.legend = None
    ws.add_chart(chart, "F27")

    ws.merge_cells("A27:D27")
    _section(ws["A27"], "How to use this workbook")
    guidance = [
        "1. Start with Prioritized Next Steps; P0/P1 access requests are the critical path.",
        "2. Filter Source Catalogue by priority, epoch role, method, access tier or verdict.",
        "3. Use Authentication Guide before creating accounts or drafting institutional requests.",
        "4. Use Nepal Pair Evidence to distinguish exact event acquisitions from generic archives.",
        "5. Do not place passwords, API keys, tokens, phone numbers or personal credentials in this file.",
    ]
    for offset, item in enumerate(guidance, 28):
        ws.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=4)
        ws.cell(offset, 1, item)
        ws.cell(offset, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(offset, 1).font = Font(name="Aptos", size=10, color=TEXT)

    for col, width in {"A": 24, "B": 26, "C": 20, "D": 22, "E": 3, "F": 25, "G": 12, "H": 12, "I": 16}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A10"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_area = "A1:I40"


def _write_next_steps(wb: Workbook) -> None:
    ws = wb.create_sheet("Prioritized Next Steps")
    ws.sheet_view.showGridLines = False
    _title(ws, "Prioritized next steps", "Ordered decision gates from data access to defensible change and burial interpretation", 10)
    headers = ["Rank", "Action", "Objective", "Why now", "Suggested owner", "Effort / lead time", "Cost / access", "Dependency", "Decision gate", "Definition of success"]
    for col, header in enumerate(headers, 1):
        ws.cell(4, col, header)
    _header(ws[4])
    for row_idx, item in enumerate(NEXT_STEPS, 5):
        for col_idx, key in enumerate(("rank", "action", "objective", "why_now", "owner", "effort", "cost_access", "dependency", "decision_gate", "success"), 1):
            ws.cell(row_idx, col_idx, item[key])
    _body(ws, 5, 4 + len(NEXT_STEPS), len(headers))
    for row_idx in range(5, 5 + len(NEXT_STEPS)):
        ws.row_dimensions[row_idx].height = 74
        rank = int(ws.cell(row_idx, 1).value)
        fill = RED if rank <= 3 else ORANGE if rank <= 7 else BLUE
        ws.cell(row_idx, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row_idx, 1).font = Font(name="Aptos", bold=True, color=TEXT, size=12)
        ws.cell(row_idx, 1).alignment = Alignment(horizontal="center", vertical="center")
    _table(ws, f"A4:J{4 + len(NEXT_STEPS)}", "PrioritizedNextSteps")
    widths = [7, 29, 31, 27, 24, 23, 23, 32, 32, 35]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{4 + len(NEXT_STEPS)}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


def _write_catalogue(wb: Workbook) -> None:
    ws = wb.create_sheet("Source Catalogue")
    ws.sheet_view.showGridLines = False
    _title(ws, "Source catalogue", "64 researched routes • exact event evidence and current entitlement are explicit • filter any column", len(SOURCE_COLUMNS))
    for col_idx, (_, header) in enumerate(SOURCE_COLUMNS, 1):
        ws.cell(4, col_idx, header)
    _header(ws[4])
    for row_idx, item in enumerate(SOURCES, 5):
        for col_idx, (key, _) in enumerate(SOURCE_COLUMNS, 1):
            ws.cell(row_idx, col_idx, item.get(key, ""))
    end_row = 4 + len(SOURCES)
    _body(ws, 5, end_row, len(SOURCE_COLUMNS))
    url_columns = {idx for idx, (key, _) in enumerate(SOURCE_COLUMNS, 1) if key.endswith("_url")}
    for row_idx in range(5, end_row + 1):
        ws.row_dimensions[row_idx].height = 88
        for col_idx in url_columns:
            _link(ws.cell(row_idx, col_idx))
        priority = ws.cell(row_idx, 2).value
        ws.cell(row_idx, 2).fill = PatternFill("solid", fgColor={"P0": RED, "P1": ORANGE, "P2": BLUE, "P3": GREY}.get(priority, WHITE))
        ws.cell(row_idx, 2).font = Font(name="Aptos", bold=True, color=TEXT)
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="center", vertical="center")
        usefulness = str(ws.cell(row_idx, 15).value).lower()
        if "very high" in usefulness or usefulness == "high":
            ws.cell(row_idx, 15).fill = PatternFill("solid", fgColor=GREEN)
        elif "low" in usefulness:
            ws.cell(row_idx, 15).fill = PatternFill("solid", fgColor=GREY)
        verdict = str(ws.cell(row_idx, 16).value).lower()
        if "reject" in verdict or "blocked" in verdict or "not" in verdict and "map" in verdict:
            ws.cell(row_idx, 16).fill = PatternFill("solid", fgColor=RED)
        access = str(ws.cell(row_idx, 18).value).lower()
        ws.cell(row_idx, 18).fill = PatternFill(
            "solid",
            fgColor=GREEN if access.startswith("open") else PURPLE if "commercial" in access else ORANGE if any(k in access for k in ("institution", "proposal", "vetted")) else BLUE,
        )
    _table(ws, f"A4:AH{end_row}", "DataSourceCatalogue")
    widths = [10, 9, 27, 22, 28, 38, 12, 14, 31, 32, 34, 34, 34, 36, 16, 35, 34, 28, 25, 38, 30, 34, 38, 34, 27, 32, 30, 36, 38, 38, 38, 12, 14, 26]
    from openpyxl.utils import get_column_letter
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "F5"
    ws.auto_filter.ref = f"A4:AH{end_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_view.zoomScale = 65


def _write_auth_guide(wb: Workbook) -> None:
    ws = wb.create_sheet("Authentication Guide")
    ws.sheet_view.showGridLines = False
    _title(ws, "Authentication and access guide", "Reusable access profiles; requirements only—never store real credentials in this workbook", 10)
    headers = ["Profile key", "Access tier", "Account required", "Signup / identity requirements", "Verification / MFA", "Eligibility / approval", "Contract / licence / payment", "API authentication", "Delivery", "Sources using profile"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(4, col_idx, header)
    _header(ws[4])
    profile_counts = Counter(s["auth_profile"] for s in SOURCES)
    rows = []
    for key, profile in AUTH_PROFILES.items():
        count = profile_counts[key]
        rows.append((key, profile, count))
    for row_idx, (key, profile, count) in enumerate(rows, 5):
        values = [key, profile["access_tier"], profile["account_required"], profile["signup_requirements"], profile["verification_mfa"], profile["eligibility_approval"], profile["contract_licence_payment"], profile["api_authentication"], profile["delivery"], count]
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, value)
    end_row = 4 + len(rows)
    _body(ws, 5, end_row, 10)
    for row_idx in range(5, end_row + 1):
        ws.row_dimensions[row_idx].height = 72
    _table(ws, f"A4:J{end_row}", "AuthenticationProfiles")
    for idx, width in enumerate([24, 30, 26, 42, 32, 38, 40, 36, 32, 16], 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:J{end_row}"
    ws.sheet_view.zoomScale = 75


def _write_pair_evidence(wb: Workbook) -> None:
    ws = wb.create_sheet("Nepal Pair Evidence")
    ws.sheet_view.showGridLines = False
    _title(ws, "Nepal acquisition and pair evidence", "The exact event candidates separated from generic provider capability", 11)
    keys = ["id", "priority", "provider", "sensor_product", "nepal_candidate", "epoch_role", "geometry", "what_exists", "provided_to_project", "verdict", "next_action"]
    headers = ["ID", "Priority", "Provider", "Sensor / product", "Exact Nepal candidate / coverage", "Epoch", "Camera / phase evidence", "What exists", "What the project has", "Decision", "Next action"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(4, col_idx, header)
    _header(ws[4])
    selected = [s for s in SOURCES if s["priority"] in {"P0", "P1"} or "Exact" in s["nepal_candidate"] or "Charter" in s["nepal_candidate"]]
    for row_idx, item in enumerate(selected, 5):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row_idx, col_idx, item[key])
    end_row = 4 + len(selected)
    _body(ws, 5, end_row, len(keys))
    for row_idx in range(5, end_row + 1):
        ws.row_dimensions[row_idx].height = 82
        priority = ws.cell(row_idx, 2).value
        ws.cell(row_idx, 2).fill = PatternFill("solid", fgColor={"P0": RED, "P1": ORANGE, "P2": BLUE, "P3": GREY}.get(priority, WHITE))
    _table(ws, f"A4:K{end_row}", "NepalPairEvidence")
    for idx, width in enumerate([10, 9, 22, 30, 43, 13, 37, 35, 37, 35, 38], 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "E5"
    ws.auto_filter.ref = f"A4:K{end_row}"
    ws.sheet_view.zoomScale = 75


def _write_definitions(wb: Workbook) -> None:
    ws = wb.create_sheet("Field Definitions")
    ws.sheet_view.showGridLines = False
    _title(ws, "Field definitions and scientific guardrails", "Interpretation rules for team members reviewing or extending the catalogue", 4)
    ws.append([])
    ws.cell(4, 1, "Field")
    ws.cell(4, 2, "Definition")
    _header(ws[4][:2])
    guardrails = FIELD_DEFINITIONS + [
        ("True topographic change", "Post-event elevation minus compatible pre-event elevation after datum harmonisation and stable-terrain co-registration, with propagated uncertainty."),
        ("Parallax residual", "Terrain-correlated displacement remaining between orthorectified views after global registration. It demonstrates a 3D signal but is not a calibrated absolute height or temporal change."),
        ("SAR requirement", "A post-event SAR image alone does not produce a terrain-change map. InSAR needs compatible complex SLC acquisitions, precise orbits, usable baseline/coherence, and masks for layover/shadow."),
        ("Resolution versus tile size", "Sub-metre imagery is the input ground sample; 1 km is only a processing/reporting tile. Output cell size is chosen from supported accuracy, not copied from image resolution."),
        ("Authentication safety", "Account requirements can be shared. Real passwords, phone numbers, MFA codes, API keys, cookies, client secrets and access tokens must stay in a local secret store or ignored environment file."),
    ]
    for row_idx, (field, definition) in enumerate(guardrails, 5):
        ws.cell(row_idx, 1, field)
        ws.cell(row_idx, 2, definition)
    end_row = 4 + len(guardrails)
    _body(ws, 5, end_row, 2)
    for row_idx in range(5, end_row + 1):
        ws.cell(row_idx, 1).font = Font(name="Aptos", bold=True, color=TEXT)
        ws.row_dimensions[row_idx].height = 48
    _table(ws, f"A4:B{end_row}", "FieldDefinitions")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 112
    ws.freeze_panes = "A5"


def _write_references(wb: Workbook) -> None:
    ws = wb.create_sheet("References")
    ws.sheet_view.showGridLines = False
    _title(ws, "Official references and access pages", "De-duplicated links used for source capability, access, signup and authentication evidence", 6)
    headers = ["Provider", "Source ID", "Reference type", "URL", "Verified date", "Host"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(4, col_idx, header)
    _header(ws[4])
    seen = set()
    refs = []
    for item in SOURCES:
        for key, label in (("source_url", "Product / programme"), ("access_url", "Access / signup"), ("auth_docs_url", "Authentication documentation")):
            url = item.get(key, "")
            if not url or url in seen:
                continue
            seen.add(url)
            refs.append((item["provider"], item["id"], label, url, item["verified_date"], urlparse(url).netloc))
    for row_idx, values in enumerate(refs, 5):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx, value)
        _link(ws.cell(row_idx, 4))
    end_row = 4 + len(refs)
    _body(ws, 5, end_row, 6)
    for row_idx in range(5, end_row + 1):
        _link(ws.cell(row_idx, 4))
        ws.row_dimensions[row_idx].height = 35
    _table(ws, f"A4:F{end_row}", "OfficialReferences")
    for col, width in zip("ABCDEF", [25, 12, 30, 95, 15, 33]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:F{end_row}"


def build_workbook(path: Path) -> None:
    generated_at = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    wb = Workbook()
    wb.properties.creator = "Nepal Flood Topographic Change Map project"
    wb.properties.title = "Nepal 2026 Flood — Topographic Data Source Catalogue"
    wb.properties.subject = "Imagery, SAR, DEM, access and authentication decision catalogue"
    wb.properties.description = "Shareable research workbook; contains no credentials or secrets."
    wb.calculation.fullCalcOnLoad = True
    _write_summary(wb, generated_at)
    _write_next_steps(wb)
    _write_catalogue(wb)
    _write_auth_guide(wb)
    _write_pair_evidence(wb)
    _write_definitions(wb)
    _write_references(wb)
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[key for key, _ in SOURCE_COLUMNS],
            lineterminator="\n",
        )
        writer.writeheader()
        for item in SOURCES:
            writer.writerow({key: item.get(key, "") for key, _ in SOURCE_COLUMNS})


def validate(xlsx_path: Path, csv_path: Path) -> dict:
    checks: list[dict[str, object]] = []

    def check(name: str, passed: bool, detail: str) -> None:
        checks.append({"name": name, "passed": bool(passed), "detail": detail})

    ids = [s["id"] for s in SOURCES]
    check("catalogue size", len(SOURCES) >= 60, f"{len(SOURCES)} sources (minimum 60)")
    check("unique source IDs", len(ids) == len(set(ids)), f"{len(set(ids))}/{len(ids)} unique")
    required = {key for key, _ in SOURCE_COLUMNS}
    missing = {s["id"]: sorted(required - s.keys()) for s in SOURCES if required - s.keys()}
    check("all required fields", not missing, json.dumps(missing) if missing else "all rows complete")
    invalid_urls = [(s["id"], key, s[key]) for s in SOURCES for key in ("source_url", "access_url", "auth_docs_url") if s.get(key) and not s[key].startswith("https://")]
    check("HTTPS references", not invalid_urls, str(invalid_urls[:5]) if invalid_urls else "all populated URLs use HTTPS")
    check("P0/P1 next actions", all(s["next_action"] for s in SOURCES if s["priority"] in {"P0", "P1"}), "all critical/high-value rows actionable")
    check("workbook exists", xlsx_path.is_file() and xlsx_path.stat().st_size > 10_000, f"{xlsx_path.stat().st_size if xlsx_path.exists() else 0} bytes")
    check("CSV exists", csv_path.is_file() and csv_path.stat().st_size > 10_000, f"{csv_path.stat().st_size if csv_path.exists() else 0} bytes")

    wb = load_workbook(xlsx_path, data_only=False, read_only=False)
    expected_sheets = ["Executive Summary", "Prioritized Next Steps", "Source Catalogue", "Authentication Guide", "Nepal Pair Evidence", "Field Definitions", "References"]
    check("workbook sheets", wb.sheetnames == expected_sheets, ", ".join(wb.sheetnames))
    ws = wb["Source Catalogue"]
    check("catalogue row count", ws.max_row == 4 + len(SOURCES), f"{ws.max_row - 4} workbook rows")
    check("catalogue filter/freeze", ws.auto_filter.ref == f"A4:AH{4 + len(SOURCES)}" and ws.freeze_panes == "F5", f"filter={ws.auto_filter.ref}, freeze={ws.freeze_panes}")
    check("catalogue Excel table", "DataSourceCatalogue" in ws.tables, f"tables={list(ws.tables)}")
    check("summary chart", len(wb["Executive Summary"]._charts) == 1, f"{len(wb['Executive Summary']._charts)} chart")
    check("reference hyperlinks", sum(bool(c.hyperlink) for c in wb["References"]["D"]) >= 40, "official source links are clickable")
    with csv_path.open(encoding="utf-8") as handle:
        csv_rows = list(csv.DictReader(handle))
    check("CSV row count", len(csv_rows) == len(SOURCES), f"{len(csv_rows)} rows")

    with zipfile.ZipFile(xlsx_path) as archive:
        workbook_text = b"\n".join(
            archive.read(name) for name in archive.namelist()
            if name.endswith((".xml", ".rels"))
        )
    payload = (workbook_text + csv_path.read_bytes()).lower()
    secret_values: list[bytes] = []
    local_env = ROOT / ".env.topographic.local"
    if local_env.exists():
        sensitive_key_fragments = ("PASS", "SECRET", "TOKEN", "KEY", "EMAIL", "PHONE", "USER", "LOGIN")
        for raw_line in local_env.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            value = value.strip().strip('"\'')
            if any(fragment in key.upper() for fragment in sensitive_key_fragments) and len(value) >= 6:
                secret_values.append(value.lower().encode())
    leaked_count = sum(value in payload for value in secret_values)
    check(
        "no local credentials",
        leaked_count == 0,
        f"{leaked_count} local secret value(s) found" if leaked_count else "no local account values embedded",
    )

    def display_path(path: Path) -> str:
        try:
            return str(path.relative_to(ROOT))
        except ValueError:
            return str(path)

    report = {
        "passed": all(c["passed"] for c in checks),
        "generated_at": datetime.now().astimezone().isoformat(),
        "source_count": len(SOURCES),
        "workbook": display_path(xlsx_path),
        "workbook_sha256": hashlib.sha256(xlsx_path.read_bytes()).hexdigest(),
        "csv": display_path(csv_path),
        "csv_sha256": hashlib.sha256(csv_path.read_bytes()).hexdigest(),
        "checks": checks,
    }
    return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    build_workbook(args.xlsx)
    write_csv(args.csv)
    report = validate(args.xlsx, args.csv)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
